#!/usr/bin/env python3
"""
offline 週報分析任務
與 offline-weekly-production-report skill 邏輯相同，以獨立腳本形式供 daemon 執行。
輸出分析摘要到 stdout，daemon 讀取後回寫 Notion。
"""

import glob
import math
import os
import shutil
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                                  Side)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: 需安裝 openpyxl → pip3 install openpyxl")
    sys.exit(1)

# ── 路徑設定 ──
NAS_DIR = Path("/Volumes/13. 公用資料夾/3. 業務相關/2. OFFLINE產品資料/安全庫存依據表")

# ── 常數 ──
TARGET_WEEKS   = 10
WEEKS_PER_MONTH = 4.33


def find_source_file():
    # 格式 YYMMDDNN（8碼，含序號），例如 offline週報分析-26062501.xlsx
    pattern = str(NAS_DIR / "offline週報分析-[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9].xlsx")
    files = sorted(glob.glob(pattern))
    if not files:
        raise FileNotFoundError(f"找不到整合來源檔於 {NAS_DIR}")
    return Path(files[-1])


def output_path():
    today = date.today()
    yy    = str(today.year)[2:]
    base  = f"offline週報分析-{yy}{today.month:02d}{today.day:02d}"
    # 自動遞增序號（01、02…），避免覆蓋當天已有的檔案
    for seq in range(1, 100):
        name = f"{base}{seq:02d}.xlsx"
        if not (NAS_DIR / name).exists():
            return NAS_DIR / name
    return NAS_DIR / f"{base}99.xlsx"


def month_range(anchor_year, anchor_month, count):
    """從 (anchor_year, anchor_month) 往前 count 個月（含當月）"""
    months = []
    y, m = anchor_year, anchor_month
    for _ in range(count):
        months.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(months))


def compute_periods():
    today = date.today()
    # 最新完整月份 = 上個月
    if today.month == 1:
        latest_y, latest_m = today.year - 1, 12
    else:
        latest_y, latest_m = today.year, today.month - 1

    months_3  = month_range(latest_y, latest_m, 3)
    months_6  = month_range(latest_y, latest_m, 6)
    months_12 = month_range(latest_y, latest_m, 12)
    return months_3, months_6, months_12


def read_inventory(wb):
    ws = wb["OFFLINE庫存數量"]
    inv = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        code = str(row[0]).strip().lower() if row[0] else ""
        name = str(row[1]).strip()         if row[1] else ""
        qty  = row[6]
        if not code or code == "none" or code == "品號":
            continue
        qty = float(qty) if isinstance(qty, (int, float)) else 0
        if code in inv:
            inv[code]["qty"] += qty
        else:
            inv[code] = {"name": name, "qty": qty}
    return inv


def read_wip(wb):
    ws = wb["OFFLINE在製數量"]
    wip = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        code = str(row[10]).strip().lower() if row[10] else ""  # col K
        name = str(row[11]).strip()          if row[11] else ""  # col L
        plan = row[12]
        done = row[13]
        if not code or code == "none" or "offline" not in code:
            continue
        plan = float(plan) if isinstance(plan, (int, float)) else 0
        done = float(done) if isinstance(done, (int, float)) else 0
        qty  = max(0, plan - done)
        if code in wip:
            wip[code]["qty"] += qty
        else:
            wip[code] = {"name": name, "qty": qty}
    return wip


def read_sales(wb):
    ws = wb["OFFLINE銷售數量"]
    sales = {}   # (year, month) → {code: qty}
    for row in ws.iter_rows(min_row=6, values_only=True):
        date_str = str(row[0]).strip() if row[0] else ""
        code     = str(row[3]).strip().lower() if row[3] else ""
        qty      = row[6]
        if len(date_str) < 6 or not code or code == "none":
            continue
        try:
            yr = int(date_str[:4])
            mo = int(date_str[4:6])
        except ValueError:
            continue
        qty = float(qty) if isinstance(qty, (int, float)) else 0
        key = (yr, mo)
        if key not in sales:
            sales[key] = {}
        sales[key][code] = sales[key].get(code, 0) + qty
    return sales


# ── Excel 樣式 ──
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def write_header_row(ws, cols, row_idx, bg, fg, bold=True):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row_idx, column=c, value=val)
        cell.fill   = make_fill(bg)
        cell.font   = Font(name="Arial", size=10, color=fg, bold=bold)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_data_cell(ws, row_idx, col_idx, val, bg=None, fg="000000", bold=False, num_fmt=None):
    cell = ws.cell(row=row_idx, column=col_idx, value=val)
    cell.font   = Font(name="Arial", size=10, color=fg, bold=bold)
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if bg:
        cell.fill = make_fill(bg)
    if num_fmt:
        cell.number_format = num_fmt


def write_sheet1(wb_out, inv, wip, sales, months_3, months_12):
    # 刪除舊分頁
    if "生產排程建議" in wb_out.sheetnames:
        del wb_out["生產排程建議"]
    ws = wb_out.create_sheet("生產排程建議", 0)

    # 標題
    month_labels = "/".join(f"{y}-{m:02d}" for y, m in months_3)
    row1_title = f"offline 生產排程建議｜近3月：{month_labels}｜目標：10週緩衝"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws.cell(row=1, column=1, value=row1_title)
    ws.cell(row=1, column=1).fill  = make_fill("1F4E79")
    ws.cell(row=1, column=1).font  = Font(name="Arial", size=11, color="FFFFFF", bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # 欄標題
    m_labels = [f"{y}-{m:02d}銷售" for y, m in months_3]
    headers = (
        ["品號", "品名", "庫存數量", "在製數量", "合計可用"]
        + m_labels
        + ["近3月合計", "月均銷量", "週均銷量", "可撐週數", "建議生產量", "生產建議"]
    )
    write_header_row(ws, headers, 2, "EBF3FB", "1F4E79")

    # 收集所有品號
    all_codes = set(inv.keys()) | set(wip.keys())

    row_idx = 3
    recommend_count = 0
    for code in sorted(all_codes):
        inv_qty  = inv.get(code, {}).get("qty", 0)
        wip_qty  = wip.get(code, {}).get("qty", 0)
        name     = inv.get(code, {}).get("name") or wip.get(code, {}).get("name", "")
        available = inv_qty + wip_qty

        month_sales = [sales.get((y, m), {}).get(code, 0) for y, m in months_3]
        total_3     = sum(month_sales)
        monthly_avg = total_3 / 3
        weekly_avg  = monthly_avg / WEEKS_PER_MONTH

        # 全年有銷
        annual_total = sum(sales.get(ym, {}).get(code, 0) for ym in months_12)

        # 過濾
        suggest_qty = max(0, math.ceil(TARGET_WEEKS * weekly_avg - available))
        if total_3 == 0 and not (annual_total > 0 and available > 0) and suggest_qty == 0:
            continue

        if weekly_avg > 0:
            holdout = round(available / weekly_avg, 1)
        else:
            holdout = None

        if suggest_qty > 0:
            suggestion = "✓ 建議安排"
            recommend_count += 1
        elif total_3 == 0 and annual_total > 0:
            suggestion = "評估中"
        else:
            suggestion = "充足"

        bg = None
        if suggestion == "✓ 建議安排":
            bg = "FFF2CC"
        elif suggestion == "充足":
            bg = None

        cells = (
            [code.upper(), name, inv_qty, wip_qty, available]
            + month_sales
            + [total_3,
               round(monthly_avg, 1),
               round(weekly_avg, 1),
               holdout,
               suggest_qty if suggest_qty > 0 else None,
               suggestion]
        )
        for c_idx, val in enumerate(cells, 1):
            last_two = len(cells) - 2
            is_qty   = c_idx == len(cells) - 1
            is_sugg  = c_idx == len(cells)
            bold_cell = (is_qty or is_sugg) and suggestion == "✓ 建議安排"
            fg_color  = "FF0000" if bold_cell else ("375623" if suggestion == "充足" and is_sugg else "000000")
            write_data_cell(ws, row_idx, c_idx, val, bg=bg, fg=fg_color, bold=bold_cell)
        row_idx += 1

    # 格式
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    col_widths = [18, 28, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return row_idx - 3, recommend_count


def write_sheet2(wb_out, inv, wip, sales, months_6, months_12):
    if "停產建議清冊" in wb_out.sheetnames:
        del wb_out["停產建議清冊"]
    ws = wb_out.create_sheet("停產建議清冊", 1)

    month_labels = "/".join(f"{y}-{m:02d}" for y, m in months_6)
    row1_title   = f"offline 停產建議清冊｜過去半年：{month_labels}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws.cell(row=1, column=1, value=row1_title)
    ws.cell(row=1, column=1).fill  = make_fill("833C00")
    ws.cell(row=1, column=1).font  = Font(name="Arial", size=11, color="FFFFFF", bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    m_labels = [f"{y}-{m:02d}" for y, m in months_6]
    headers  = (
        ["品號", "品名", "庫存數量", "在製數量", "合計可用"]
        + m_labels
        + ["半年合計", "全年合計", "建議"]
    )
    write_header_row(ws, headers, 2, "FCE4D6", "833C00")

    all_codes = set(inv.keys()) | set(wip.keys())

    row_idx  = 3
    has_stock_count = 0

    for code in sorted(all_codes):
        half_sales = [sales.get((y, m), {}).get(code, 0) for y, m in months_6]
        if sum(half_sales) > 0:
            continue   # 半年有銷，不納入

        inv_qty   = inv.get(code, {}).get("qty", 0)
        wip_qty   = wip.get(code, {}).get("qty", 0)
        name      = inv.get(code, {}).get("name") or wip.get(code, {}).get("name", "")
        available = inv_qty + wip_qty

        annual = sum(sales.get(ym, {}).get(code, 0) for ym in months_12)

        has_stock = available > 0
        if has_stock:
            has_stock_count += 1

        bg_stock = "FFF2CC" if has_stock else None
        cells = (
            [code.upper(), name, inv_qty, wip_qty, available]
            + half_sales
            + [0, round(annual, 0), "⚠ 停產評估"]
        )
        for c_idx, val in enumerate(cells, 1):
            is_sugg = c_idx == len(cells)
            stock_cols = {3, 4, 5}
            bg = bg_stock if c_idx in stock_cols else ("FCE4D6" if is_sugg else None)
            write_data_cell(ws, row_idx, c_idx, val, bg=bg,
                            fg="FF0000" if is_sugg else "000000",
                            bold=is_sugg)
        row_idx += 1

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    col_widths = [18, 28, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return row_idx - 3, has_stock_count


def main():
    source = find_source_file()
    output = output_path()

    # 複製來源檔
    if source != output:
        shutil.copy2(source, output)

    wb_out = openpyxl.load_workbook(output)
    wb_src = wb_out  # 同一個檔案

    months_3, months_6, months_12 = compute_periods()

    inv   = read_inventory(wb_src)
    wip   = read_wip(wb_src)
    sales = read_sales(wb_src)

    s1_total, s1_recommend = write_sheet1(wb_out, inv, wip, sales, months_3, months_12)
    s2_total, s2_stock     = write_sheet2(wb_out, inv, wip, sales, months_6, months_12)

    wb_out.save(output)

    # 輸出摘要（daemon 擷取後寫入 Notion 結果摘要欄）
    m3 = "/".join(f"{y}-{m:02d}" for y, m in months_3)
    m6 = "/".join(f"{y}-{m:02d}" for y, m in months_6)
    print(
        f"來源：{source.name} → 輸出：{output.name}\n"
        f"近3月：{m3}｜半年：{m6[:15]}...\n"
        f"生產建議：共{s1_total}筆，建議安排{s1_recommend}筆\n"
        f"停產清冊：共{s2_total}筆，仍有庫存{s2_stock}筆"
    )


if __name__ == "__main__":
    main()
